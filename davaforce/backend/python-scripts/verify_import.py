from __future__ import annotations

import argparse
import math
import sqlite3
import sys
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_PATH = ROOT_DIR / "sample_data" /"input_data_sample.xlsx"
DEFAULT_DB_PATH = ROOT_DIR / "prisma" / "workforce_sample.db"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Verify that the SQLite database matches the workforce planning workbook."
    )
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH)
    return parser.parse_args()


def text(value: Any) -> str:
    if value is None:
        return ""
    try:
        return value.isoformat()
    except AttributeError:
        return str(value).strip()


def as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def as_bool(value: Any) -> bool:
    cleaned = text(value).lower()
    if cleaned in {"yes", "true", "1"}:
        return True
    if cleaned in {"no", "false", "0", ""}:
        return False
    raise ValueError(f"Cannot convert {value!r} to bool.")


def as_date(value: Any) -> date:
    if isinstance(value, date):
        return value
    return date.fromisoformat(text(value))


def non_empty_rows(sheet: Any) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    output: list[dict[str, Any]] = []
    for row in rows[1:]:
        if any(value is not None and str(value).strip() != "" for value in row):
            output.append(dict(zip(header, row)))
    return output


def semicolon_item_count(value: Any) -> int:
    return len([item.strip() for item in text(value).split(";") if item.strip()])


def sqlite_count(connection: sqlite3.Connection, table_name: str) -> int:
    row = connection.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()
    return int(row[0])


def add_result(results: list[tuple[bool, str, str]], passed: bool, name: str, detail: str) -> None:
    results.append((passed, name, detail))


def compare_scalar(expected: Any, actual: Any) -> bool:
    if isinstance(expected, (int, float)) or isinstance(actual, (int, float)):
        try:
            return math.isclose(float(expected), float(actual), rel_tol=0.0, abs_tol=1e-9)
        except (TypeError, ValueError):
            return False
    return text(expected) == text(actual)


def first_full_week_on_or_after(release_date: date) -> date:
    return release_date + timedelta(days=(7 - release_date.weekday()) % 7)


def verify_counts(
    workbook_rows: dict[str, list[dict[str, Any]]], connection: sqlite3.Connection
) -> list[tuple[bool, str, str]]:
    results: list[tuple[bool, str, str]] = []

    table_expectations = [
        ("Person", len(workbook_rows["People"])),
        ("PersonAvailabilitySnapshot", len(workbook_rows["People"])),
        ("Profile", len(workbook_rows["Profiles"])),
        ("SkillCatalog", len(workbook_rows["Skill Catalog"])),
        ("PersonSkillEvidence", len(workbook_rows["Skills"])),
        ("CurrentAllocation", len(workbook_rows["Allocations"])),
        ("SupplyRecord", len(workbook_rows["Bench"])),
        ("PartialCapacityView", len(workbook_rows["Partial Capacity"])),
        ("AvailabilityWeek", len(workbook_rows["Availability Calendar"])),
        ("BenchMovementWeek", len(workbook_rows["Bench Movement"])),
        ("ProjectHistory", len(workbook_rows["Project History"])),
        ("Opportunity", len(workbook_rows["Opportunities"])),
        ("OpportunityRole", len(workbook_rows["Opportunity Roles"])),
        ("OpportunityCandidateOverlay", len(workbook_rows["Opportunity Overlays"])),
        ("EwaRequest", len(workbook_rows["EWA Requests"])),
        ("ScenarioTarget", len(workbook_rows["Scenario Targets"])),
    ]

    role_skill_count = 0
    for row in workbook_rows["Opportunity Roles"]:
        role_skill_count += semicolon_item_count(row["RequiredSkills"])
        role_skill_count += semicolon_item_count(row["DesiredSkills"])
    table_expectations.append(("OpportunityRoleSkillRequirement", role_skill_count))

    raw_row_count = sum(len(rows) for rows in workbook_rows.values())
    table_expectations.append(("RawSheetRow", raw_row_count))

    for table_name, expected_count in table_expectations:
        actual_count = sqlite_count(connection, table_name)
        add_result(
            results,
            actual_count == expected_count,
            f"count:{table_name}",
            f"expected={expected_count} actual={actual_count}",
        )

    return results


def verify_integrity(connection: sqlite3.Connection) -> list[tuple[bool, str, str]]:
    results: list[tuple[bool, str, str]] = []
    queries = {
        "profiles_without_person": """
            SELECT COUNT(*) FROM "Profile" p
            LEFT JOIN "Person" x ON x.id = p.personId
            WHERE x.id IS NULL
        """,
        "skills_without_person": """
            SELECT COUNT(*) FROM "PersonSkillEvidence" s
            LEFT JOIN "Person" p ON p.id = s.personId
            WHERE p.id IS NULL
        """,
        "skills_without_catalog": """
            SELECT COUNT(*) FROM "PersonSkillEvidence" s
            LEFT JOIN "SkillCatalog" c ON c.name = s.skillName
            WHERE c.name IS NULL
        """,
        "allocations_without_person": """
            SELECT COUNT(*) FROM "CurrentAllocation" a
            LEFT JOIN "Person" p ON p.id = a.personId
            WHERE p.id IS NULL
        """,
        "supply_without_person": """
            SELECT COUNT(*) FROM "SupplyRecord" s
            LEFT JOIN "Person" p ON p.id = s.personId
            WHERE p.id IS NULL
        """,
        "partial_capacity_without_person": """
            SELECT COUNT(*) FROM "PartialCapacityView" v
            LEFT JOIN "Person" p ON p.id = v.personId
            WHERE p.id IS NULL
        """,
        "partial_capacity_missing_source": """
            SELECT COUNT(*) FROM "PartialCapacityView" v
            LEFT JOIN "SupplyRecord" s ON s.id = v.sourceBenchRecordId
            WHERE s.id IS NULL
        """,
        "availability_without_person": """
            SELECT COUNT(*) FROM "AvailabilityWeek" a
            LEFT JOIN "Person" p ON p.id = a.personId
            WHERE p.id IS NULL
        """,
        "history_without_person": """
            SELECT COUNT(*) FROM "ProjectHistory" h
            LEFT JOIN "Person" p ON p.id = h.personId
            WHERE p.id IS NULL
        """,
        "roles_without_opportunity": """
            SELECT COUNT(*) FROM "OpportunityRole" r
            LEFT JOIN "Opportunity" o ON o.id = r.opportunityId
            WHERE o.id IS NULL
        """,
        "role_skills_without_role": """
            SELECT COUNT(*) FROM "OpportunityRoleSkillRequirement" rs
            LEFT JOIN "OpportunityRole" r ON r.id = rs.opportunityRoleId
            WHERE r.id IS NULL
        """,
        "role_skills_without_catalog": """
            SELECT COUNT(*) FROM "OpportunityRoleSkillRequirement" rs
            LEFT JOIN "SkillCatalog" s ON s.name = rs.skillName
            WHERE s.name IS NULL
        """,
        "overlays_without_opportunity": """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay" o
            LEFT JOIN "Opportunity" p ON p.id = o.opportunityId
            WHERE p.id IS NULL
        """,
        "overlays_without_role": """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay" o
            LEFT JOIN "OpportunityRole" r ON r.id = o.opportunityRoleId
            WHERE r.id IS NULL
        """,
        "overlays_without_person": """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay" o
            LEFT JOIN "Person" p ON p.id = o.personId
            WHERE p.id IS NULL
        """,
        "ewa_without_opportunity": """
            SELECT COUNT(*) FROM "EwaRequest" e
            LEFT JOIN "Opportunity" o ON o.id = e.opportunityId
            WHERE o.id IS NULL
        """,
        "ewa_without_role": """
            SELECT COUNT(*) FROM "EwaRequest" e
            LEFT JOIN "OpportunityRole" r ON r.id = e.opportunityRoleId
            WHERE r.id IS NULL
        """,
        "ewa_without_person": """
            SELECT COUNT(*) FROM "EwaRequest" e
            LEFT JOIN "Person" p ON p.id = e.personId
            WHERE p.id IS NULL
        """,
        "duplicate_person_skill": """
            SELECT COUNT(*) FROM (
              SELECT personId, skillName, COUNT(*) c
              FROM "PersonSkillEvidence"
              GROUP BY personId, skillName
              HAVING c > 1
            )
        """,
        "duplicate_availability_week": """
            SELECT COUNT(*) FROM (
              SELECT personId, weekStartDate, COUNT(*) c
              FROM "AvailabilityWeek"
              GROUP BY personId, weekStartDate
              HAVING c > 1
            )
        """,
        "duplicate_overlay_role_person": """
            SELECT COUNT(*) FROM (
              SELECT opportunityRoleId, personId, COUNT(*) c
              FROM "OpportunityCandidateOverlay"
              GROUP BY opportunityRoleId, personId
              HAVING c > 1
            )
        """,
        "duplicate_ewa_role_person": """
            SELECT COUNT(*) FROM (
              SELECT opportunityRoleId, personId, COUNT(*) c
              FROM "EwaRequest"
              GROUP BY opportunityRoleId, personId
              HAVING c > 1
            )
        """,
        "overlay_ewa_status_mismatch": """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay" o
            JOIN "EwaRequest" e
              ON e.opportunityRoleId = o.opportunityRoleId
             AND e.personId = o.personId
            WHERE o.ewaStatus <> e.ewaStatus
        """,
        "overlay_matchscore_mismatch": """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay"
            WHERE ABS(matchScore - capabilityFitScore) > 1e-9
        """,
    }

    for name, sql in queries.items():
        actual = int(connection.execute(sql).fetchone()[0])
        add_result(results, actual == 0, f"integrity:{name}", f"actual={actual}")

    return results


def verify_spot_checks(
    workbook_rows: dict[str, list[dict[str, Any]]], connection: sqlite3.Connection
) -> list[tuple[bool, str, str]]:
    results: list[tuple[bool, str, str]] = []

    people_row = workbook_rows["People"][0]
    person = connection.execute(
        """
        SELECT p.name, s.availabilityCategory, s.availableFteCurrent, s.expectedReleaseDate
        FROM "Person" p
        JOIN "PersonAvailabilitySnapshot" s ON s.personId = p.id
        WHERE p.id = ?
        """,
        (text(people_row["Employee_ID"]),),
    ).fetchone()
    add_result(
        results,
        person is not None
        and compare_scalar(people_row["Employee_Name"], person[0])
        and compare_scalar(people_row["AvailabilityCategory"], person[1])
        and compare_scalar(people_row["AvailableFTECurrent"], person[2])
        and compare_scalar(people_row["ExpectedReleaseDate"], person[3]),
        "spot:people_first_row",
        f"employee_id={text(people_row['Employee_ID'])}",
    )

    skill_row = workbook_rows["Skills"][0]
    skill = connection.execute(
        """
        SELECT skillName, skillLevel, yearsExperience, confidence
        FROM "PersonSkillEvidence"
        WHERE id = ?
        """,
        (text(skill_row["Skill_Row_ID"]),),
    ).fetchone()
    add_result(
        results,
        skill is not None
        and compare_scalar(skill_row["SkillName"], skill[0])
        and compare_scalar(skill_row["SkillLevel"], skill[1])
        and compare_scalar(skill_row["YearsExperience"], skill[2])
        and compare_scalar(skill_row["Confidence"], skill[3]),
        "spot:skills_first_row",
        f"skill_row_id={text(skill_row['Skill_Row_ID'])}",
    )

    allocation_row = workbook_rows["Allocations"][0]
    allocation = connection.execute(
        """
        SELECT accountId, clientName, clientType, projectId, projectName, domain, allocationFte
        FROM "CurrentAllocation"
        WHERE id = ?
        """,
        (text(allocation_row["Allocation_ID"]),),
    ).fetchone()
    add_result(
        results,
        allocation is not None
        and compare_scalar(allocation_row["AccountID"], allocation[0])
        and compare_scalar(allocation_row["Client_Name"], allocation[1])
        and compare_scalar(allocation_row["Client_Type"], allocation[2])
        and compare_scalar(allocation_row["ProjectID"], allocation[3])
        and compare_scalar(allocation_row["Project_Name"], allocation[4])
        and compare_scalar(allocation_row["Domain"], allocation[5])
        and compare_scalar(allocation_row["AllocationFTE"], allocation[6]),
        "spot:allocations_first_row",
        f"allocation_id={text(allocation_row['Allocation_ID'])}",
    )

    role_row = workbook_rows["Opportunity Roles"][0]
    role = connection.execute(
        """
        SELECT requiredSkillsText, desiredSkillsText, fteRequired, canCombineCandidates
        FROM "OpportunityRole"
        WHERE id = ?
        """,
        (text(role_row["Opportunity_Role_ID"]),),
    ).fetchone()
    add_result(
        results,
        role is not None
        and compare_scalar(role_row["RequiredSkills"], role[0])
        and compare_scalar(role_row["DesiredSkills"], role[1])
        and compare_scalar(role_row["FTERequired"], role[2])
        and compare_scalar(1 if text(role_row["CanCombineCandidates"]).lower() == "yes" else 0, role[3]),
        "spot:opportunity_role_first_row",
        f"opportunity_role_id={text(role_row['Opportunity_Role_ID'])}",
    )

    overlay_row = workbook_rows["Opportunity Overlays"][0]
    overlay = connection.execute(
        """
        SELECT fitStatus, rank, matchScore, ewaStatus
        FROM "OpportunityCandidateOverlay"
        WHERE id = ?
        """,
        (text(overlay_row["Overlay_ID"]),),
    ).fetchone()
    add_result(
        results,
        overlay is not None
        and compare_scalar(overlay_row["FitStatus"], overlay[0])
        and compare_scalar(overlay_row["Rank"], overlay[1])
        and compare_scalar(overlay_row["MatchScore"], overlay[2])
        and compare_scalar(overlay_row["EWAStatus"], overlay[3]),
        "spot:overlay_first_row",
        f"overlay_id={text(overlay_row['Overlay_ID'])}",
    )

    ewa_row = workbook_rows["EWA Requests"][0]
    ewa = connection.execute(
        """
        SELECT requestType, ewaStatus, requestedFte, nextAction
        FROM "EwaRequest"
        WHERE id = ?
        """,
        (text(ewa_row["EWA_Request_ID"]),),
    ).fetchone()
    add_result(
        results,
        ewa is not None
        and compare_scalar(ewa_row["RequestType"], ewa[0])
        and compare_scalar(ewa_row["EWAStatus"], ewa[1])
        and compare_scalar(ewa_row["RequestedFTE"], ewa[2])
        and compare_scalar(ewa_row["NextAction"], ewa[3]),
        "spot:ewa_first_row",
        f"ewa_request_id={text(ewa_row['EWA_Request_ID'])}",
    )

    scenario_row = workbook_rows["Scenario Targets"][0]
    scenario = connection.execute(
        """
        SELECT scenarioName, targetDate, targetBenchRate, targetBenchHeadcount
        FROM "ScenarioTarget"
        WHERE id = ?
        """,
        (text(scenario_row["Scenario_ID"]),),
    ).fetchone()
    add_result(
        results,
        scenario is not None
        and compare_scalar(scenario_row["ScenarioName"], scenario[0])
        and compare_scalar(scenario_row["TargetDate"], scenario[1])
        and compare_scalar(scenario_row["TargetBenchRate"], scenario[2])
        and compare_scalar(scenario_row["TargetBenchHeadcount"], scenario[3]),
        "spot:scenario_first_row",
        f"scenario_id={text(scenario_row['Scenario_ID'])}",
    )

    return results


def verify_validation_summary_criteria(
    workbook_rows: dict[str, list[dict[str, Any]]], connection: sqlite3.Connection
) -> list[tuple[bool, str, str]]:
    results: list[tuple[bool, str, str]] = []

    people = workbook_rows["People"]
    skills = workbook_rows["Skills"]
    bench = workbook_rows["Bench"]
    partial = workbook_rows["Partial Capacity"]
    availability = workbook_rows["Availability Calendar"]
    history = workbook_rows["Project History"]
    opportunities = workbook_rows["Opportunities"]
    roles = workbook_rows["Opportunity Roles"]
    overlays = workbook_rows["Opportunity Overlays"]
    ewa_requests = workbook_rows["EWA Requests"]

    people_by_id = {text(row["Employee_ID"]): row for row in people}
    role_by_id = {text(row["Opportunity_Role_ID"]): row for row in roles}
    history_people = {text(row["Employee_ID"]) for row in history}
    person_grades = {text(row["Grade"]) for row in people}

    skill_people_by_name: dict[str, set[str]] = defaultdict(set)
    for row in skills:
        skill_people_by_name[text(row["SkillName"])].add(text(row["Employee_ID"]))

    availability_by_person: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in availability:
        availability_by_person[text(row["Employee_ID"])].append(row)
    for rows in availability_by_person.values():
        rows.sort(key=lambda item: as_date(item["WeekStartDate"]))

    role_ranks: dict[str, set[int]] = defaultdict(set)
    for row in overlays:
        role_ranks[text(row["Opportunity_Role_ID"])].add(int(float(row["Rank"])))

    validation_rows = {
        text(row["Check_ID"]): row for row in workbook_rows.get("Validation Summary", [])
    }

    def record(check_id: str, computed: Any, passed: bool, detail: str) -> None:
        workbook_actual = (
            text(validation_rows[check_id]["Actual"]) if check_id in validation_rows else "n/a"
        )
        add_result(
            results,
            passed,
            f"validation:{check_id}",
            f"computed={computed} workbook_actual={workbook_actual}; {detail}",
        )

    # VAL-001 .. VAL-007
    val_001 = sqlite_count(connection, "Person")
    record("VAL-001", val_001, val_001 == len(people), "employee count")

    val_002 = int(
        connection.execute(
            'SELECT COUNT(*) FROM "SupplyRecord" WHERE supplyType = ?', ("Current Bench",)
        ).fetchone()[0]
    )
    expected_002 = sum(1 for row in bench if text(row["BenchType"]) == "Current Bench")
    record("VAL-002", val_002, val_002 == expected_002, "current bench count")

    val_003 = int(
        connection.execute(
            'SELECT COUNT(*) FROM "SupplyRecord" WHERE supplyType = ?', ("Partial Capacity",)
        ).fetchone()[0]
    )
    expected_003 = sum(1 for row in bench if text(row["BenchType"]) == "Partial Capacity")
    record("VAL-003", val_003, val_003 == expected_003, "partial-capacity count")

    val_004 = int(
        connection.execute(
            'SELECT COUNT(*) FROM "SupplyRecord" WHERE supplyType = ?', ("Future Roll-off",)
        ).fetchone()[0]
    )
    expected_004 = sum(1 for row in bench if text(row["BenchType"]) == "Future Roll-off")
    record("VAL-004", val_004, val_004 == expected_004, "future roll-off count")

    val_005 = sqlite_count(connection, "Opportunity")
    record("VAL-005", val_005, val_005 == len(opportunities), "opportunity count")

    val_006 = sqlite_count(connection, "OpportunityRole")
    record("VAL-006", val_006, val_006 == len(roles), "opportunity-role count")

    val_007 = sqlite_count(connection, "AvailabilityWeek")
    expected_007 = len(people) * 12
    record("VAL-007", val_007, val_007 == expected_007, "availability rows")

    # VAL-008 duplicate primary IDs in workbook source data
    primary_key_columns = {
        "People": "Employee_ID",
        "Skills": "Skill_Row_ID",
        "Profiles": "Profile_ID",
        "Allocations": "Allocation_ID",
        "Bench": "Bench_Record_ID",
        "Partial Capacity": "Bench_Record_ID",
        "Availability Calendar": "Availability_ID",
        "Bench Movement": "WeekStartDate",
        "Project History": "History_ID",
        "Opportunities": "Opportunity_ID",
        "Opportunity Roles": "Opportunity_Role_ID",
        "Opportunity Overlays": "Overlay_ID",
        "EWA Requests": "EWA_Request_ID",
        "Scenario Targets": "Scenario_ID",
    }
    duplicate_primary_ids = 0
    for sheet_name, key_column in primary_key_columns.items():
        values = [text(row[key_column]) for row in workbook_rows[sheet_name]]
        duplicate_primary_ids += len(values) - len(set(values))
    record("VAL-008", duplicate_primary_ids, duplicate_primary_ids == 0, "duplicate primary IDs")

    # VAL-009 foreign-key errors
    foreign_key_errors = sum(
        int(connection.execute(sql).fetchone()[0])
        for sql in [
            """
            SELECT COUNT(*) FROM "Profile" p
            LEFT JOIN "Person" x ON x.id = p.personId
            WHERE x.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "PersonSkillEvidence" s
            LEFT JOIN "Person" p ON p.id = s.personId
            WHERE p.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "PersonSkillEvidence" s
            LEFT JOIN "SkillCatalog" c ON c.name = s.skillName
            WHERE c.name IS NULL
            """,
            """
            SELECT COUNT(*) FROM "CurrentAllocation" a
            LEFT JOIN "Person" p ON p.id = a.personId
            WHERE p.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "SupplyRecord" s
            LEFT JOIN "Person" p ON p.id = s.personId
            WHERE p.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "PartialCapacityView" v
            LEFT JOIN "Person" p ON p.id = v.personId
            WHERE p.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "PartialCapacityView" v
            LEFT JOIN "SupplyRecord" s ON s.id = v.sourceBenchRecordId
            WHERE s.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "AvailabilityWeek" a
            LEFT JOIN "Person" p ON p.id = a.personId
            WHERE p.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "ProjectHistory" h
            LEFT JOIN "Person" p ON p.id = h.personId
            WHERE p.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "OpportunityRole" r
            LEFT JOIN "Opportunity" o ON o.id = r.opportunityId
            WHERE o.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "OpportunityRoleSkillRequirement" rs
            LEFT JOIN "OpportunityRole" r ON r.id = rs.opportunityRoleId
            WHERE r.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "OpportunityRoleSkillRequirement" rs
            LEFT JOIN "SkillCatalog" s ON s.name = rs.skillName
            WHERE s.name IS NULL
            """,
            """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay" o
            LEFT JOIN "Opportunity" p ON p.id = o.opportunityId
            WHERE p.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay" o
            LEFT JOIN "OpportunityRole" r ON r.id = o.opportunityRoleId
            WHERE r.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay" o
            LEFT JOIN "Person" p ON p.id = o.personId
            WHERE p.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "EwaRequest" e
            LEFT JOIN "Opportunity" o ON o.id = e.opportunityId
            WHERE o.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "EwaRequest" e
            LEFT JOIN "OpportunityRole" r ON r.id = e.opportunityRoleId
            WHERE r.id IS NULL
            """,
            """
            SELECT COUNT(*) FROM "EwaRequest" e
            LEFT JOIN "Person" p ON p.id = e.personId
            WHERE p.id IS NULL
            """,
        ]
    )
    record("VAL-009", foreign_key_errors, foreign_key_errors == 0, "foreign-key errors")

    # VAL-010 employee name mismatches
    name_mismatch_count = 0
    for sheet_name in [
        "Skills",
        "Profiles",
        "Allocations",
        "Bench",
        "Partial Capacity",
        "Availability Calendar",
        "Project History",
        "Opportunity Overlays",
        "EWA Requests",
    ]:
        for row in workbook_rows[sheet_name]:
            employee_id = text(row["Employee_ID"])
            if text(row["Employee_Name"]) != text(people_by_id[employee_id]["Employee_Name"]):
                name_mismatch_count += 1
    record("VAL-010", name_mismatch_count, name_mismatch_count == 0, "employee-name mismatches")

    # VAL-011 overlay / EWA status mismatches
    val_011 = int(
        connection.execute(
            """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay" o
            JOIN "EwaRequest" e
              ON e.opportunityRoleId = o.opportunityRoleId
             AND e.personId = o.personId
            WHERE o.ewaStatus <> e.ewaStatus
            """
        ).fetchone()[0]
    )
    record("VAL-011", val_011, val_011 == 0, "overlay/EWA status mismatches")

    # VAL-012 FTE gaps without blocking reason
    val_012 = int(
        connection.execute(
            """
            SELECT COUNT(*) FROM "EwaRequest"
            WHERE fteGap > 0
              AND TRIM(COALESCE(blockingReason, '')) IN ('', 'None')
            """
        ).fetchone()[0]
    )
    record("VAL-012", val_012, val_012 == 0, "FTE gaps without blocking reason")

    # VAL-013 overlay candidates without project history
    val_013 = int(
        connection.execute(
            """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay" o
            LEFT JOIN "ProjectHistory" h ON h.personId = o.personId
            WHERE h.personId IS NULL
            """
        ).fetchone()[0]
    )
    record("VAL-013", val_013, val_013 == 0, "overlay candidates without project history")

    # VAL-014 invalid grade preferences
    val_014 = sum(1 for row in roles if text(row["GradePreference"]) not in person_grades)
    record("VAL-014", val_014, val_014 == 0, "invalid grade preferences")

    # VAL-015 domain labels left in DesiredSkills
    domain_labels = {
        text(row["PrimaryDomain"]) for row in people
    } | {
        text(row["SecondaryDomain"]) for row in people
    } | {
        text(row["Domain"]) for row in opportunities
    }
    val_015 = 0
    for row in roles:
        for skill in [item.strip() for item in text(row["DesiredSkills"]).split(";") if item.strip()]:
            if skill in domain_labels:
                val_015 += 1
    record("VAL-015", val_015, val_015 == 0, "domain labels left in DesiredSkills")

    # VAL-016 intentionally unsupported required skill names
    unsupported_required_skills = sorted(
        {
            skill.strip()
            for row in roles
            for skill in text(row["RequiredSkills"]).split(";")
            if skill.strip() and len(skill_people_by_name.get(skill.strip(), set())) == 0
        }
    )
    val_016 = len(unsupported_required_skills)
    record(
        "VAL-016",
        val_016,
        val_016 == 3,
        f"unsupported required skills={', '.join(unsupported_required_skills)}",
    )

    # VAL-017 future roll-offs available before full planning week
    val_017 = 0
    for row in bench:
        if text(row["BenchType"]) != "Future Roll-off":
            continue
        employee_id = text(row["Employee_ID"])
        release_date = as_date(people_by_id[employee_id]["ExpectedReleaseDate"])
        threshold_week = first_full_week_on_or_after(release_date)
        first_positive_week = next(
            (
                as_date(week["WeekStartDate"])
                for week in availability_by_person[employee_id]
                if as_float(week["AvailableFTE"]) > 0
            ),
            None,
        )
        if first_positive_week is not None and first_positive_week < threshold_week:
            val_017 += 1
    record(
        "VAL-017",
        val_017,
        val_017 == 0,
        "future roll-offs available before first full planning week",
    )

    # VAL-018 partial-capacity weekly transition errors
    val_018 = 0
    for row in people:
        if text(row["AvailabilityCategory"]) != "Partial Capacity":
            continue
        employee_id = text(row["Employee_ID"])
        current_fte = as_float(row["AvailableFTECurrent"])
        release_date = as_date(row["ExpectedReleaseDate"])
        transition_week = first_full_week_on_or_after(release_date)
        for week in availability_by_person[employee_id]:
            week_start = as_date(week["WeekStartDate"])
            expected_fte = 1.0 if week_start >= transition_week else current_fte
            if not math.isclose(
                as_float(week["AvailableFTE"]), expected_fte, rel_tol=0.0, abs_tol=1e-9
            ):
                val_018 += 1
                break
    record("VAL-018", val_018, val_018 == 0, "partial-capacity weekly transition errors")

    # VAL-019 employees without exactly 12 calendar rows
    val_019 = sum(
        1 for _, count in Counter(text(row["Employee_ID"]) for row in availability).items() if count != 12
    )
    record("VAL-019", val_019, val_019 == 0, "employees without exactly 12 calendar rows")

    # VAL-020 EWA role/date/FTE structural errors
    val_020 = 0
    overlay_by_pair = {
        (text(row["Opportunity_Role_ID"]), text(row["Employee_ID"])): row for row in overlays
    }
    for row in ewa_requests:
        role = role_by_id[text(row["Opportunity_Role_ID"])]
        overlay = overlay_by_pair.get((text(row["Opportunity_Role_ID"]), text(row["Employee_ID"])))

        if text(row["Opportunity_ID"]) != text(role["Opportunity_ID"]):
            val_020 += 1
            continue
        if as_date(row["ProposedStartDate"]) != as_date(role["StartDate"]):
            val_020 += 1
            continue
        expected_end_date = as_date(role["StartDate"]) + timedelta(days=int(as_float(role["DurationWeeks"])) * 7)
        if as_date(row["ProposedEndDate"]) != expected_end_date:
            val_020 += 1
            continue
        if as_float(row["RequestedFTE"]) <= 0 or as_float(row["RequestedFTE"]) > as_float(role["FTERequired"]):
            val_020 += 1
            continue
        if as_float(row["RequestedFTE"]) < as_float(role["MinimumIndividualFTE"]):
            val_020 += 1
            continue
        if as_bool(row["CanSplitRole"]) != as_bool(role["CanCombineCandidates"]):
            val_020 += 1
            continue
        if not as_bool(role["CanCombineCandidates"]) and not math.isclose(
            as_float(row["RequestedFTE"]),
            as_float(role["FTERequired"]),
            rel_tol=0.0,
            abs_tol=1e-9,
        ):
            val_020 += 1
            continue
        if overlay is not None:
            if not math.isclose(
                as_float(row["AvailableFTEAtStart"]),
                as_float(overlay["AvailableFTEAtStart"]),
                rel_tol=0.0,
                abs_tol=1e-9,
            ):
                val_020 += 1
                continue
            if not math.isclose(
                as_float(row["FTEGap"]),
                as_float(overlay["FTEGap"]),
                rel_tol=0.0,
                abs_tol=1e-9,
            ):
                val_020 += 1
                continue
            if as_date(row["EarliestFullAvailabilityDate"]) != as_date(
                overlay["EarliestFullAvailabilityDate"]
            ):
                val_020 += 1
                continue
    record("VAL-020", val_020, val_020 == 0, "EWA role/date/FTE structural errors")

    # VAL-021 bench / partial-capacity duplicate flags incorrect
    partial_people = {text(row["Employee_ID"]) for row in partial}
    val_021 = 0
    for row in bench:
        expected_flag = text(row["Employee_ID"]) in partial_people
        if as_bool(row["IsAlsoInPartialCapacityView"]) != expected_flag:
            val_021 += 1
    record("VAL-021", val_021, val_021 == 0, "bench/partial-capacity duplicate flags incorrect")

    # VAL-022 out-of-range match scores
    val_022 = int(
        connection.execute(
            """
            SELECT COUNT(*) FROM "OpportunityCandidateOverlay"
            WHERE matchScore < 0 OR matchScore > 100
               OR capabilityFitScore < 0 OR capabilityFitScore > 100
               OR availabilityFitScore < 0 OR availabilityFitScore > 100
               OR overallStaffingScore < 0 OR overallStaffingScore > 100
            """
        ).fetchone()[0]
    )
    record("VAL-022", val_022, val_022 == 0, "out-of-range match scores")

    # VAL-023 roles without ranks 1,2,3
    val_023 = 0
    for row in roles:
        if role_ranks[text(row["Opportunity_Role_ID"])] != {1, 2, 3}:
            val_023 += 1
    record("VAL-023", val_023, val_023 == 0, "roles without ranks 1,2,3")

    # VAL-025 intentional-gap roles must be blocked if requested
    intentional_gap_skill_names = set(unsupported_required_skills)
    intentional_gap_roles = {
        text(row["Opportunity_Role_ID"])
        for row in roles
        if any(
            skill.strip() in intentional_gap_skill_names
            for skill in text(row["RequiredSkills"]).split(";")
            if skill.strip()
        )
    }
    val_025 = 0
    for row in ewa_requests:
        if text(row["Opportunity_Role_ID"]) not in intentional_gap_roles:
            continue
        if text(row["EWAStatus"]) != "Blocked":
            val_025 += 1
            continue
        if text(row["BlockingReason"]) in {"", "None"}:
            val_025 += 1
    record(
        "VAL-025",
        val_025,
        val_025 == 0,
        "EWA requests for intentional-gap roles not blocked",
    )

    return results


def main() -> None:
    args = parse_args()
    excel_path = args.excel.resolve()
    db_path = args.db.resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if not db_path.exists():
        raise FileNotFoundError(f"SQLite database not found: {db_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    workbook_rows = {sheet_name: non_empty_rows(workbook[sheet_name]) for sheet_name in workbook.sheetnames}

    connection = sqlite3.connect(db_path)
    try:
        results: list[tuple[bool, str, str]] = []
        results.extend(verify_counts(workbook_rows, connection))
        results.extend(verify_integrity(connection))
        results.extend(verify_spot_checks(workbook_rows, connection))
        results.extend(verify_validation_summary_criteria(workbook_rows, connection))
    finally:
        connection.close()

    passed = 0
    failed = 0
    for ok, name, detail in results:
        status = "PASS" if ok else "FAIL"
        print(f"{status} | {name} | {detail}")
        if ok:
            passed += 1
        else:
            failed += 1

    print(
        "\nNote: VAL-024 (10,000-iteration portfolio stress test) is not re-executed here "
        "because it is a scenario simulation, not a persisted-data normalization/integrity check."
    )
    print(f"\nSummary: passed={passed} failed={failed}")
    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
