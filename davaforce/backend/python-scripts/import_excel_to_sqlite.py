from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_PATH = ROOT_DIR / "sample_data" /"input_data_sample.xlsx"
DEFAULT_DB_PATH = ROOT_DIR / "prisma" / "workforce_sample.db"

REQUIRED_CANONICAL_SHEETS = [
    "People",
    "Skills",
    "Skill Catalog",
    "Profiles",
    "Allocations",
    "Bench",
    "Partial Capacity",
    "Availability Calendar",
    "Bench Movement",
    "Project History",
    "Opportunities",
    "Opportunity Roles",
    "Opportunity Overlays",
    "EWA Requests",
    "Scenario Targets",
]

NATURAL_KEY_COLUMNS = {
    "README": "Item",
    "Dataset Summary": "Metric",
    "Data Dictionary": "Column",
    "People": "Employee_ID",
    "Skills": "Skill_Row_ID",
    "Skill Catalog": "SkillName",
    "Profiles": "Profile_ID",
    "Allocations": "Allocation_ID",
    "Bench": "Bench_Record_ID",
    "Partial Capacity": "Bench_Record_ID",
    "Availability Calendar": "Availability_ID",
    "Bench Movement": "WeekStartDate",
    "Project History": "History_ID",
    "Opportunities": "Opportunity_ID",
    "Opportunity Roles": "Opportunity_Role_ID",
    "Opportunity Overlays": "Overlay_ID",
    "EWA Requests": "EWA_Request_ID",
    "Scenario Targets": "Scenario_ID",
    "Starter Prompts": "Prompt_ID",
    "Change Log": "Change_ID",
    "Validation Summary": "Check_ID",
}

SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE "ImportBatch" (
  "id" INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
  "workbookName" TEXT NOT NULL,
  "workbookVersion" TEXT,
  "importedAt" TEXT NOT NULL
);

CREATE TABLE "RawSheetRow" (
  "id" INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
  "importBatchId" INTEGER NOT NULL,
  "sheetName" TEXT NOT NULL,
  "sourceRowNumber" INTEGER NOT NULL,
  "naturalKey" TEXT,
  "rowHash" TEXT NOT NULL,
  "payloadJson" TEXT NOT NULL,
  FOREIGN KEY ("importBatchId") REFERENCES "ImportBatch"("id") ON DELETE CASCADE
);

CREATE UNIQUE INDEX "RawSheetRow_importBatch_sheet_row_key"
  ON "RawSheetRow"("importBatchId", "sheetName", "sourceRowNumber");
CREATE INDEX "RawSheetRow_sheetName_idx" ON "RawSheetRow"("sheetName");
CREATE INDEX "RawSheetRow_naturalKey_idx" ON "RawSheetRow"("naturalKey");

CREATE TABLE "Person" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "name" TEXT NOT NULL,
  "region" TEXT NOT NULL,
  "country" TEXT NOT NULL,
  "city" TEXT NOT NULL,
  "timezone" TEXT NOT NULL,
  "department" TEXT NOT NULL,
  "discipline" TEXT NOT NULL,
  "roleArchetype" TEXT NOT NULL,
  "grade" TEXT NOT NULL,
  "careerLevel" INTEGER NOT NULL,
  "primaryDomain" TEXT NOT NULL,
  "secondaryDomain" TEXT NOT NULL,
  "workMode" TEXT NOT NULL
);

CREATE TABLE "PersonAvailabilitySnapshot" (
  "personId" TEXT NOT NULL PRIMARY KEY,
  "availabilityCategory" TEXT NOT NULL,
  "currentAllocationFte" REAL NOT NULL,
  "availableFteCurrent" REAL NOT NULL,
  "expectedReleaseDate" TEXT NOT NULL,
  "releaseWindow" TEXT NOT NULL,
  "ewaStatus" TEXT NOT NULL,
  "currentAccountId" TEXT,
  "currentProjectId" TEXT,
  "currentRole" TEXT,
  "currentProjectStart" TEXT,
  "currentProjectEnd" TEXT,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE
);

CREATE INDEX "PersonAvailabilitySnapshot_category_idx"
  ON "PersonAvailabilitySnapshot"("availabilityCategory");
CREATE INDEX "PersonAvailabilitySnapshot_release_idx"
  ON "PersonAvailabilitySnapshot"("expectedReleaseDate");

CREATE TABLE "Profile" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "personId" TEXT NOT NULL UNIQUE,
  "profileSummary" TEXT NOT NULL,
  "keyStrengthsText" TEXT NOT NULL,
  "preferredWorkTypes" TEXT NOT NULL,
  "domainExperienceSummary" TEXT NOT NULL,
  "certificationsText" TEXT NOT NULL,
  "recentHighlights" TEXT NOT NULL,
  "mobilityNotes" TEXT NOT NULL,
  "languagesText" TEXT NOT NULL,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE
);

CREATE TABLE "SkillCatalog" (
  "name" TEXT NOT NULL PRIMARY KEY,
  "category" TEXT NOT NULL,
  "description" TEXT NOT NULL,
  "relevantDepartmentsText" TEXT NOT NULL,
  "suggestedLevelScaleText" TEXT NOT NULL
);

CREATE INDEX "SkillCatalog_category_idx" ON "SkillCatalog"("category");

CREATE TABLE "PersonSkillEvidence" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "personId" TEXT NOT NULL,
  "skillName" TEXT NOT NULL,
  "skillLevel" INTEGER NOT NULL,
  "yearsExperience" REAL NOT NULL,
  "lastUsedDate" TEXT NOT NULL,
  "evidenceSource" TEXT NOT NULL,
  "confidence" TEXT NOT NULL,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE,
  FOREIGN KEY ("skillName") REFERENCES "SkillCatalog"("name")
);

CREATE UNIQUE INDEX "PersonSkillEvidence_person_skill_key"
  ON "PersonSkillEvidence"("personId", "skillName");
CREATE INDEX "PersonSkillEvidence_skill_idx" ON "PersonSkillEvidence"("skillName");
CREATE INDEX "PersonSkillEvidence_confidence_idx" ON "PersonSkillEvidence"("confidence");

CREATE TABLE "CurrentAllocation" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "personId" TEXT NOT NULL UNIQUE,
  "accountId" TEXT NOT NULL,
  "clientName" TEXT NOT NULL,
  "clientType" TEXT NOT NULL,
  "projectId" TEXT NOT NULL,
  "projectName" TEXT NOT NULL,
  "domain" TEXT NOT NULL,
  "roleOnProject" TEXT NOT NULL,
  "allocationFte" REAL NOT NULL,
  "startDate" TEXT NOT NULL,
  "plannedEndDate" TEXT NOT NULL,
  "allocationStatus" TEXT NOT NULL,
  "ewaStatus" TEXT NOT NULL,
  "lastUpdated" TEXT NOT NULL,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE
);

CREATE INDEX "CurrentAllocation_account_idx" ON "CurrentAllocation"("accountId");
CREATE INDEX "CurrentAllocation_project_idx" ON "CurrentAllocation"("projectId");
CREATE INDEX "CurrentAllocation_end_date_idx" ON "CurrentAllocation"("plannedEndDate");
CREATE INDEX "CurrentAllocation_status_idx" ON "CurrentAllocation"("allocationStatus");

CREATE TABLE "SupplyRecord" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "personId" TEXT NOT NULL UNIQUE,
  "supplyType" TEXT NOT NULL,
  "availabilityCategory" TEXT NOT NULL,
  "availableFrom" TEXT NOT NULL,
  "supplyFte" REAL NOT NULL,
  "supplyPercent" REAL NOT NULL,
  "primaryDomain" TEXT NOT NULL,
  "topSkillsText" TEXT NOT NULL,
  "supplyRisk" TEXT NOT NULL,
  "timeOnSupplyDays" INTEGER NOT NULL,
  "suggestedAction" TEXT NOT NULL,
  "targetRoleFit" TEXT NOT NULL,
  "ewaActionRequired" TEXT NOT NULL,
  "isAlsoInPartialCapacityView" INTEGER NOT NULL,
  "recordUsage" TEXT NOT NULL,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE
);

CREATE INDEX "SupplyRecord_type_idx" ON "SupplyRecord"("supplyType");
CREATE INDEX "SupplyRecord_category_idx" ON "SupplyRecord"("availabilityCategory");
CREATE INDEX "SupplyRecord_available_from_idx" ON "SupplyRecord"("availableFrom");
"""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a SQLite database from the workforce planning workbook."
    )
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Replace the target SQLite file if it already exists.",
    )
    return parser.parse_args()


def read_sheet(workbook: Any, sheet_name: str) -> dict[str, Any]:
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {"header": [], "rows": []}

    header = [str(value) if value is not None else "" for value in rows[0]]
    result_rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value is not None and str(value).strip() != "" for value in row):
            continue
        values = dict(zip(header, row))
        result_rows.append({"row_number": row_number, "values": values})
    return {"header": header, "rows": result_rows}


def normalize_for_json(value: Any) -> Any:
    if isinstance(value, datetime):
        if value.time() == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def row_payload_json(row: dict[str, Any]) -> str:
    normalized = {key: normalize_for_json(value) for key, value in row.items()}
    return json.dumps(normalized, ensure_ascii=True, sort_keys=True)


def row_hash(payload_json: str) -> str:
    return hashlib.sha256(payload_json.encode("utf-8")).hexdigest()


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def optional_text(value: Any) -> str | None:
    cleaned = text(value)
    return cleaned or None


def as_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    return int(float(value))


def as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    cleaned = text(value).lower()
    if cleaned in {"yes", "true", "1"}:
        return True
    if cleaned in {"no", "false", "0", ""}:
        return False
    raise ValueError(f"Cannot coerce {value!r} to bool.")


def parse_semicolon_list(value: Any) -> list[str]:
    items = [item.strip() for item in text(value).split(";")]
    return [item for item in items if item]


def make_client_id(name: str, client_type: str) -> str:
    raw = f"{name.lower()}|{client_type.lower()}"
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return f"CLI-{digest.upper()}"


def workbook_version(sheets: dict[str, dict[str, Any]]) -> str | None:
    if "README" not in sheets:
        return None
    readme_rows = sheets["README"]["rows"]
    metadata = {
        text(row["values"].get("Item")): text(row["values"].get("Details"))
        for row in readme_rows
    }
    return metadata.get("Version") or None


def ensure_required_sheets(sheets: dict[str, dict[str, Any]]) -> None:
    missing = [sheet for sheet in REQUIRED_CANONICAL_SHEETS if sheet not in sheets]
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(missing)}")


def create_database(db_path: Path, replace: bool) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        if not replace:
            raise FileExistsError(
                f"{db_path} already exists. Re-run with --replace to rebuild it."
            )
        db_path.unlink()
    connection = sqlite3.connect(db_path)
    connection.execute("PRAGMA foreign_keys = ON")
    connection.executescript(SCHEMA_SQL)
    return connection


EXTRA_SCHEMA_SQL = """
CREATE TABLE "PartialCapacityView" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "personId" TEXT NOT NULL UNIQUE,
  "sourceBenchRecordId" TEXT NOT NULL,
  "benchType" TEXT NOT NULL,
  "availabilityCategory" TEXT NOT NULL,
  "availableFrom" TEXT NOT NULL,
  "benchFte" REAL NOT NULL,
  "benchPercent" REAL NOT NULL,
  "primaryDomain" TEXT NOT NULL,
  "topSkillsText" TEXT NOT NULL,
  "benchRisk" TEXT NOT NULL,
  "timeOnBenchDays" INTEGER NOT NULL,
  "suggestedAction" TEXT NOT NULL,
  "targetRoleFit" TEXT NOT NULL,
  "ewaActionRequired" TEXT NOT NULL,
  "viewType" TEXT NOT NULL,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE,
  FOREIGN KEY ("sourceBenchRecordId") REFERENCES "SupplyRecord"("id")
);

CREATE INDEX "PartialCapacityView_source_idx"
  ON "PartialCapacityView"("sourceBenchRecordId");

CREATE TABLE "AvailabilityWeek" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "personId" TEXT NOT NULL,
  "weekStartDate" TEXT NOT NULL,
  "availableFte" REAL NOT NULL,
  "availabilityType" TEXT NOT NULL,
  "source" TEXT NOT NULL,
  "confidence" TEXT NOT NULL,
  "ewaStatus" TEXT NOT NULL,
  "notes" TEXT NOT NULL,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE
);

CREATE UNIQUE INDEX "AvailabilityWeek_person_week_key"
  ON "AvailabilityWeek"("personId", "weekStartDate");
CREATE INDEX "AvailabilityWeek_week_idx" ON "AvailabilityWeek"("weekStartDate");
CREATE INDEX "AvailabilityWeek_type_idx" ON "AvailabilityWeek"("availabilityType");

CREATE TABLE "BenchMovementWeek" (
  "weekStartDate" TEXT NOT NULL PRIMARY KEY,
  "currentBenchHeadcount" INTEGER NOT NULL,
  "emergingBenchHeadcount" INTEGER NOT NULL,
  "partialCapacityHeadcount" INTEGER NOT NULL,
  "availableFte" REAL NOT NULL,
  "notes" TEXT NOT NULL
);

CREATE TABLE "ProjectHistory" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "personId" TEXT NOT NULL,
  "clientName" TEXT NOT NULL,
  "clientType" TEXT NOT NULL,
  "projectName" TEXT NOT NULL,
  "domain" TEXT NOT NULL,
  "role" TEXT NOT NULL,
  "startDate" TEXT NOT NULL,
  "endDate" TEXT NOT NULL,
  "keyTechnologiesOrMethods" TEXT NOT NULL,
  "responsibilities" TEXT NOT NULL,
  "outcomeEvidence" TEXT NOT NULL,
  "region" TEXT NOT NULL,
  "teamSize" INTEGER NOT NULL,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE
);

CREATE INDEX "ProjectHistory_person_idx" ON "ProjectHistory"("personId");
CREATE INDEX "ProjectHistory_domain_idx" ON "ProjectHistory"("domain");
CREATE INDEX "ProjectHistory_end_date_idx" ON "ProjectHistory"("endDate");

CREATE TABLE "Opportunity" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "clientName" TEXT NOT NULL,
  "clientType" TEXT NOT NULL,
  "name" TEXT NOT NULL,
  "region" TEXT NOT NULL,
  "country" TEXT NOT NULL,
  "city" TEXT NOT NULL,
  "domain" TEXT NOT NULL,
  "stage" TEXT NOT NULL,
  "probability" REAL NOT NULL,
  "expectedStartDate" TEXT NOT NULL,
  "durationWeeks" INTEGER NOT NULL,
  "commercialPriority" TEXT NOT NULL,
  "deliveryRisk" TEXT NOT NULL,
  "opportunityBrief" TEXT NOT NULL,
  "timezonePreference" TEXT NOT NULL
);

CREATE INDEX "Opportunity_client_idx" ON "Opportunity"("clientName");
CREATE INDEX "Opportunity_stage_idx" ON "Opportunity"("stage");
CREATE INDEX "Opportunity_start_idx" ON "Opportunity"("expectedStartDate");

CREATE TABLE "OpportunityRole" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "opportunityId" TEXT NOT NULL,
  "roleName" TEXT NOT NULL,
  "disciplineOrDepartment" TEXT NOT NULL,
  "gradePreference" TEXT NOT NULL,
  "requiredSkillsText" TEXT NOT NULL,
  "desiredSkillsText" TEXT NOT NULL,
  "domainExperienceRequired" TEXT NOT NULL,
  "locationPreference" TEXT NOT NULL,
  "startDate" TEXT NOT NULL,
  "durationWeeks" INTEGER NOT NULL,
  "fteRequired" REAL NOT NULL,
  "priority" TEXT NOT NULL,
  "flexibilityNotes" TEXT NOT NULL,
  "minimumIndividualFte" REAL NOT NULL,
  "canCombineCandidates" INTEGER NOT NULL,
  FOREIGN KEY ("opportunityId") REFERENCES "Opportunity"("id") ON DELETE CASCADE
);

CREATE INDEX "OpportunityRole_opp_idx" ON "OpportunityRole"("opportunityId");
CREATE INDEX "OpportunityRole_start_idx" ON "OpportunityRole"("startDate");
CREATE INDEX "OpportunityRole_priority_idx" ON "OpportunityRole"("priority");

CREATE TABLE "OpportunityRoleSkillRequirement" (
  "opportunityRoleId" TEXT NOT NULL,
  "skillName" TEXT NOT NULL,
  "importance" TEXT NOT NULL,
  PRIMARY KEY ("opportunityRoleId", "skillName", "importance"),
  FOREIGN KEY ("opportunityRoleId") REFERENCES "OpportunityRole"("id") ON DELETE CASCADE,
  FOREIGN KEY ("skillName") REFERENCES "SkillCatalog"("name")
);

CREATE INDEX "OpportunityRoleSkillRequirement_skill_idx"
  ON "OpportunityRoleSkillRequirement"("skillName");
CREATE INDEX "OpportunityRoleSkillRequirement_importance_idx"
  ON "OpportunityRoleSkillRequirement"("importance");
"""


def extend_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(EXTRA_SCHEMA_SQL)


def insert_import_batch(
    connection: sqlite3.Connection, workbook_name: str, version: str | None
) -> int:
    cursor = connection.execute(
        """
        INSERT INTO "ImportBatch" ("workbookName", "workbookVersion", "importedAt")
        VALUES (?, ?, ?)
        """,
        (
            workbook_name,
            version,
            datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        ),
    )
    return int(cursor.lastrowid)


def load_raw_sheet_rows(
    connection: sqlite3.Connection,
    import_batch_id: int,
    sheets: dict[str, dict[str, Any]],
) -> int:
    rows_to_insert: list[tuple[Any, ...]] = []
    for sheet_name, sheet_data in sheets.items():
        natural_key_column = NATURAL_KEY_COLUMNS.get(sheet_name)
        for row in sheet_data["rows"]:
            values = row["values"]
            payload_json = row_payload_json(values)
            natural_key = (
                optional_text(values.get(natural_key_column)) if natural_key_column else None
            )
            rows_to_insert.append(
                (
                    import_batch_id,
                    sheet_name,
                    row["row_number"],
                    natural_key,
                    row_hash(payload_json),
                    payload_json,
                )
            )
    connection.executemany(
        """
        INSERT INTO "RawSheetRow"
          ("importBatchId", "sheetName", "sourceRowNumber", "naturalKey", "rowHash", "payloadJson")
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        rows_to_insert,
    )
    return len(rows_to_insert)


def insert_people(connection: sqlite3.Connection, rows: list[dict[str, Any]]) -> int:
    payload = [
        (
            text(row["Employee_ID"]),
            text(row["Employee_Name"]),
            text(row["Region"]),
            text(row["Country"]),
            text(row["City"]),
            text(row["Timezone"]),
            text(row["Department"]),
            text(row["Discipline"]),
            text(row["RoleArchetype"]),
            text(row["Grade"]),
            as_int(row["CareerLevel"]),
            text(row["PrimaryDomain"]),
            text(row["SecondaryDomain"]),
            text(row["WorkMode"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "Person"
          ("id", "name", "region", "country", "city", "timezone", "department",
           "discipline", "roleArchetype", "grade", "careerLevel", "primaryDomain",
           "secondaryDomain", "workMode")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_person_snapshots(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["Employee_ID"]),
            text(row["AvailabilityCategory"]),
            as_float(row["CurrentAllocationFTE"]),
            as_float(row["AvailableFTECurrent"]),
            text(row["ExpectedReleaseDate"]),
            text(row["ReleaseWindow"]),
            text(row["EWAStatus"]),
            optional_text(row["CurrentAccountID"]),
            optional_text(row["CurrentProjectID"]),
            optional_text(row["CurrentRole"]),
            optional_text(row["CurrentProjectStart"]),
            optional_text(row["CurrentProjectEnd"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "PersonAvailabilitySnapshot"
          ("personId", "availabilityCategory", "currentAllocationFte", "availableFteCurrent",
           "expectedReleaseDate", "releaseWindow", "ewaStatus", "currentAccountId",
           "currentProjectId", "currentRole", "currentProjectStart", "currentProjectEnd")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


FINAL_SCHEMA_SQL = """
CREATE TABLE "OpportunityCandidateOverlay" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "opportunityId" TEXT NOT NULL,
  "opportunityRoleId" TEXT NOT NULL,
  "personId" TEXT NOT NULL,
  "fitStatus" TEXT NOT NULL,
  "rank" INTEGER NOT NULL,
  "matchScore" REAL NOT NULL,
  "rationale" TEXT NOT NULL,
  "constraint" TEXT NOT NULL,
  "ewaStatus" TEXT NOT NULL,
  "plannerNotes" TEXT NOT NULL,
  "capabilityFitScore" REAL NOT NULL,
  "availabilityFitScore" REAL NOT NULL,
  "overallStaffingScore" REAL NOT NULL,
  "availableFteAtStart" REAL NOT NULL,
  "fteGap" REAL NOT NULL,
  "earliestFullAvailabilityDate" TEXT NOT NULL,
  "requiredSkillsMatched" INTEGER NOT NULL,
  "requiredSkillsTotal" INTEGER NOT NULL,
  "desiredSkillsMatched" INTEGER NOT NULL,
  "desiredSkillsTotal" INTEGER NOT NULL,
  FOREIGN KEY ("opportunityId") REFERENCES "Opportunity"("id") ON DELETE CASCADE,
  FOREIGN KEY ("opportunityRoleId") REFERENCES "OpportunityRole"("id") ON DELETE CASCADE,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE
);

CREATE UNIQUE INDEX "OpportunityCandidateOverlay_role_person_key"
  ON "OpportunityCandidateOverlay"("opportunityRoleId", "personId");
CREATE INDEX "OpportunityCandidateOverlay_opp_idx"
  ON "OpportunityCandidateOverlay"("opportunityId");
CREATE INDEX "OpportunityCandidateOverlay_person_idx"
  ON "OpportunityCandidateOverlay"("personId");
CREATE INDEX "OpportunityCandidateOverlay_rank_idx"
  ON "OpportunityCandidateOverlay"("rank");

CREATE TABLE "EwaRequest" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "opportunityId" TEXT NOT NULL,
  "opportunityRoleId" TEXT NOT NULL,
  "personId" TEXT NOT NULL,
  "requestType" TEXT NOT NULL,
  "ewaStatus" TEXT NOT NULL,
  "requestedFte" REAL NOT NULL,
  "proposedStartDate" TEXT NOT NULL,
  "proposedEndDate" TEXT NOT NULL,
  "approvalRequired" INTEGER NOT NULL,
  "bookingOwner" TEXT NOT NULL,
  "blockingReason" TEXT NOT NULL,
  "nextAction" TEXT NOT NULL,
  "lastUpdated" TEXT NOT NULL,
  "notes" TEXT NOT NULL,
  "availableFteAtStart" REAL NOT NULL,
  "fteGap" REAL NOT NULL,
  "canSplitRole" INTEGER NOT NULL,
  "earliestFullAvailabilityDate" TEXT NOT NULL,
  FOREIGN KEY ("opportunityId") REFERENCES "Opportunity"("id") ON DELETE CASCADE,
  FOREIGN KEY ("opportunityRoleId") REFERENCES "OpportunityRole"("id") ON DELETE CASCADE,
  FOREIGN KEY ("personId") REFERENCES "Person"("id") ON DELETE CASCADE
);

CREATE UNIQUE INDEX "EwaRequest_role_person_key"
  ON "EwaRequest"("opportunityRoleId", "personId");
CREATE INDEX "EwaRequest_opp_idx" ON "EwaRequest"("opportunityId");
CREATE INDEX "EwaRequest_person_idx" ON "EwaRequest"("personId");
CREATE INDEX "EwaRequest_status_idx" ON "EwaRequest"("ewaStatus");

CREATE TABLE "ScenarioTarget" (
  "id" TEXT NOT NULL PRIMARY KEY,
  "scenarioName" TEXT NOT NULL,
  "targetDate" TEXT NOT NULL,
  "targetBenchRate" REAL NOT NULL,
  "targetBenchHeadcount" INTEGER NOT NULL,
  "focus" TEXT NOT NULL,
  "successMeasure" TEXT NOT NULL
);

CREATE INDEX "ScenarioTarget_date_idx" ON "ScenarioTarget"("targetDate");
"""


def finalize_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(FINAL_SCHEMA_SQL)


def insert_profiles(connection: sqlite3.Connection, rows: list[dict[str, Any]]) -> int:
    payload = [
        (
            text(row["Profile_ID"]),
            text(row["Employee_ID"]),
            text(row["ProfileSummary"]),
            text(row["KeyStrengths"]),
            text(row["PreferredWorkTypes"]),
            text(row["DomainExperienceSummary"]),
            text(row["Certifications"]),
            text(row["RecentHighlights"]),
            text(row["MobilityNotes"]),
            text(row["Languages"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "Profile"
          ("id", "personId", "profileSummary", "keyStrengthsText", "preferredWorkTypes",
           "domainExperienceSummary", "certificationsText", "recentHighlights",
           "mobilityNotes", "languagesText")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_skill_catalog(connection: sqlite3.Connection, rows: list[dict[str, Any]]) -> int:
    payload = [
        (
            text(row["SkillName"]),
            text(row["SkillCategory"]),
            text(row["Description"]),
            text(row["RelevantDepartments"]),
            text(row["SuggestedLevelScale"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "SkillCatalog"
          ("name", "category", "description", "relevantDepartmentsText", "suggestedLevelScaleText")
        VALUES (?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_person_skills(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["Skill_Row_ID"]),
            text(row["Employee_ID"]),
            text(row["SkillName"]),
            as_int(row["SkillLevel"]),
            as_float(row["YearsExperience"]),
            text(row["LastUsedDate"]),
            text(row["EvidenceSource"]),
            text(row["Confidence"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "PersonSkillEvidence"
          ("id", "personId", "skillName", "skillLevel", "yearsExperience",
           "lastUsedDate", "evidenceSource", "confidence")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_clients(connection: sqlite3.Connection, sheets: dict[str, dict[str, Any]]) -> dict[tuple[str, str], str]:
    pairs: set[tuple[str, str]] = set()
    for row in sheets["Allocations"]["rows"]:
        values = row["values"]
        pairs.add((text(values["Client_Name"]), text(values["Client_Type"])))
    for row in sheets["Opportunities"]["rows"]:
        values = row["values"]
        pairs.add((text(values["Client_Name"]), text(values["Client_Type"])))

    payload = []
    mapping: dict[tuple[str, str], str] = {}
    for name, client_type in sorted(pairs):
        client_id = make_client_id(name, client_type)
        mapping[(name, client_type)] = client_id
        payload.append((client_id, name, client_type))

    connection.executemany(
        """
        INSERT INTO "Client" ("id", "name", "clientType")
        VALUES (?, ?, ?)
        """,
        payload,
    )
    return mapping


def insert_accounts_and_projects(
    connection: sqlite3.Connection,
    allocation_rows: list[dict[str, Any]],
    client_ids: dict[tuple[str, str], str],
) -> tuple[int, int]:
    account_payload: dict[str, tuple[str, str]] = {}
    project_payload: dict[str, tuple[str, str, str]] = {}

    for row in allocation_rows:
        client_key = (text(row["Client_Name"]), text(row["Client_Type"]))
        client_id = client_ids[client_key]
        account_id = text(row["AccountID"])
        project_id = text(row["ProjectID"])

        existing_account = account_payload.get(account_id)
        if existing_account and existing_account != (account_id, client_id):
            raise ValueError(f"Account {account_id} maps to multiple clients.")
        account_payload[account_id] = (account_id, client_id)

        project_tuple = (
            project_id,
            account_id,
            text(row["Project_Name"]),
            text(row["Domain"]),
        )
        existing_project = project_payload.get(project_id)
        if existing_project and existing_project != project_tuple:
            raise ValueError(f"Project {project_id} has conflicting definitions.")
        project_payload[project_id] = project_tuple

    connection.executemany(
        """
        INSERT INTO "Account" ("id", "clientId")
        VALUES (?, ?)
        """,
        sorted(account_payload.values()),
    )
    connection.executemany(
        """
        INSERT INTO "Project" ("id", "accountId", "name", "domain")
        VALUES (?, ?, ?, ?)
        """,
        sorted(project_payload.values()),
    )
    return len(account_payload), len(project_payload)


def insert_current_allocations(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["Allocation_ID"]),
            text(row["Employee_ID"]),
            text(row["AccountID"]),
            text(row["Client_Name"]),
            text(row["Client_Type"]),
            text(row["ProjectID"]),
            text(row["Project_Name"]),
            text(row["Domain"]),
            text(row["RoleOnProject"]),
            as_float(row["AllocationFTE"]),
            text(row["StartDate"]),
            text(row["PlannedEndDate"]),
            text(row["AllocationStatus"]),
            text(row["EWAStatus"]),
            text(row["LastUpdated"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "CurrentAllocation"
          ("id", "personId", "accountId", "clientName", "clientType", "projectId",
           "projectName", "domain", "roleOnProject", "allocationFte", "startDate",
           "plannedEndDate", "allocationStatus", "ewaStatus", "lastUpdated")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_supply_records(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["Bench_Record_ID"]),
            text(row["Employee_ID"]),
            text(row["BenchType"]),
            text(row["AvailabilityCategory"]),
            text(row["AvailableFrom"]),
            as_float(row["BenchFTE"]),
            as_float(row["BenchPercent"]),
            text(row["PrimaryDomain"]),
            text(row["TopSkills"]),
            text(row["BenchRisk"]),
            as_int(row["TimeOnBenchDays"]),
            text(row["SuggestedAction"]),
            text(row["TargetRoleFit"]),
            text(row["EWAActionRequired"]),
            as_bool(row["IsAlsoInPartialCapacityView"]),
            text(row["RecordUsage"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "SupplyRecord"
          ("id", "personId", "supplyType", "availabilityCategory", "availableFrom",
           "supplyFte", "supplyPercent", "primaryDomain", "topSkillsText", "supplyRisk",
           "timeOnSupplyDays", "suggestedAction", "targetRoleFit", "ewaActionRequired",
           "isAlsoInPartialCapacityView", "recordUsage")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_partial_capacity_view(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["Bench_Record_ID"]),
            text(row["Employee_ID"]),
            text(row["SourceBenchRecordID"]),
            text(row["BenchType"]),
            text(row["AvailabilityCategory"]),
            text(row["AvailableFrom"]),
            as_float(row["BenchFTE"]),
            as_float(row["BenchPercent"]),
            text(row["PrimaryDomain"]),
            text(row["TopSkills"]),
            text(row["BenchRisk"]),
            as_int(row["TimeOnBenchDays"]),
            text(row["SuggestedAction"]),
            text(row["TargetRoleFit"]),
            text(row["EWAActionRequired"]),
            text(row["ViewType"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "PartialCapacityView"
          ("id", "personId", "sourceBenchRecordId", "benchType", "availabilityCategory",
           "availableFrom", "benchFte", "benchPercent", "primaryDomain", "topSkillsText",
           "benchRisk", "timeOnBenchDays", "suggestedAction", "targetRoleFit",
           "ewaActionRequired", "viewType")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_availability_weeks(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["Availability_ID"]),
            text(row["Employee_ID"]),
            text(row["WeekStartDate"]),
            as_float(row["AvailableFTE"]),
            text(row["AvailabilityType"]),
            text(row["Source"]),
            text(row["Confidence"]),
            text(row["EWAStatus"]),
            text(row["Notes"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "AvailabilityWeek"
          ("id", "personId", "weekStartDate", "availableFte", "availabilityType",
           "source", "confidence", "ewaStatus", "notes")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_bench_movement(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["WeekStartDate"]),
            as_int(row["CurrentBenchHeadcount"]),
            as_int(row["EmergingBenchHeadcount"]),
            as_int(row["PartialCapacityHeadcount"]),
            as_float(row["AvailableFTE"]),
            text(row["Notes"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "BenchMovementWeek"
          ("weekStartDate", "currentBenchHeadcount", "emergingBenchHeadcount",
           "partialCapacityHeadcount", "availableFte", "notes")
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_project_history(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["History_ID"]),
            text(row["Employee_ID"]),
            text(row["Client_Name"]),
            text(row["Client_Type"]),
            text(row["Project_Name"]),
            text(row["Domain"]),
            text(row["Role"]),
            text(row["StartDate"]),
            text(row["EndDate"]),
            text(row["KeyTechnologiesOrMethods"]),
            text(row["Responsibilities"]),
            text(row["OutcomeEvidence"]),
            text(row["Region"]),
            as_int(row["TeamSize"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "ProjectHistory"
          ("id", "personId", "clientName", "clientType", "projectName", "domain", "role",
           "startDate", "endDate", "keyTechnologiesOrMethods", "responsibilities",
           "outcomeEvidence", "region", "teamSize")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_opportunities(
    connection: sqlite3.Connection,
    rows: list[dict[str, Any]],
) -> int:
    payload = [
        (
            text(row["Opportunity_ID"]),
            text(row["Client_Name"]),
            text(row["Client_Type"]),
            text(row["Opportunity_Name"]),
            text(row["Region"]),
            text(row["Country"]),
            text(row["City"]),
            text(row["Domain"]),
            text(row["Stage"]),
            as_float(row["Probability"]),
            text(row["ExpectedStartDate"]),
            as_int(row["DurationWeeks"]),
            text(row["CommercialPriority"]),
            text(row["DeliveryRisk"]),
            text(row["OpportunityBrief"]),
            text(row["TimezonePreference"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "Opportunity"
          ("id", "clientName", "clientType", "name", "region", "country", "city", "domain", "stage",
           "probability", "expectedStartDate", "durationWeeks", "commercialPriority",
           "deliveryRisk", "opportunityBrief", "timezonePreference")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_opportunity_roles(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> tuple[int, int]:
    role_payload = [
        (
            text(row["Opportunity_Role_ID"]),
            text(row["Opportunity_ID"]),
            text(row["RoleName"]),
            text(row["DisciplineOrDepartment"]),
            text(row["GradePreference"]),
            text(row["RequiredSkills"]),
            text(row["DesiredSkills"]),
            text(row["DomainExperienceRequired"]),
            text(row["LocationPreference"]),
            text(row["StartDate"]),
            as_int(row["DurationWeeks"]),
            as_float(row["FTERequired"]),
            text(row["Priority"]),
            text(row["FlexibilityNotes"]),
            as_float(row["MinimumIndividualFTE"]),
            as_bool(row["CanCombineCandidates"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "OpportunityRole"
          ("id", "opportunityId", "roleName", "disciplineOrDepartment", "gradePreference",
           "requiredSkillsText", "desiredSkillsText", "domainExperienceRequired",
           "locationPreference", "startDate", "durationWeeks", "fteRequired", "priority",
           "flexibilityNotes", "minimumIndividualFte", "canCombineCandidates")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        role_payload,
    )

    requirement_payload: list[tuple[str, str, str]] = []
    for row in rows:
        role_id = text(row["Opportunity_Role_ID"])
        for skill_name in parse_semicolon_list(row["RequiredSkills"]):
            requirement_payload.append((role_id, skill_name, "REQUIRED"))
        for skill_name in parse_semicolon_list(row["DesiredSkills"]):
            requirement_payload.append((role_id, skill_name, "DESIRED"))

    connection.executemany(
        """
        INSERT INTO "OpportunityRoleSkillRequirement"
          ("opportunityRoleId", "skillName", "importance")
        VALUES (?, ?, ?)
        """,
        requirement_payload,
    )
    return len(role_payload), len(requirement_payload)


def insert_overlays(connection: sqlite3.Connection, rows: list[dict[str, Any]]) -> int:
    payload = [
        (
            text(row["Overlay_ID"]),
            text(row["Opportunity_ID"]),
            text(row["Opportunity_Role_ID"]),
            text(row["Employee_ID"]),
            text(row["FitStatus"]),
            as_int(row["Rank"]),
            as_float(row["MatchScore"]),
            text(row["Rationale"]),
            text(row["Constraint"]),
            text(row["EWAStatus"]),
            text(row["PlannerNotes"]),
            as_float(row["CapabilityFitScore"]),
            as_float(row["AvailabilityFitScore"]),
            as_float(row["OverallStaffingScore"]),
            as_float(row["AvailableFTEAtStart"]),
            as_float(row["FTEGap"]),
            text(row["EarliestFullAvailabilityDate"]),
            as_int(row["RequiredSkillsMatched"]),
            as_int(row["RequiredSkillsTotal"]),
            as_int(row["DesiredSkillsMatched"]),
            as_int(row["DesiredSkillsTotal"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "OpportunityCandidateOverlay"
          ("id", "opportunityId", "opportunityRoleId", "personId", "fitStatus", "rank",
           "matchScore", "rationale", "constraint", "ewaStatus", "plannerNotes",
           "capabilityFitScore", "availabilityFitScore", "overallStaffingScore",
           "availableFteAtStart", "fteGap", "earliestFullAvailabilityDate",
           "requiredSkillsMatched", "requiredSkillsTotal", "desiredSkillsMatched",
           "desiredSkillsTotal")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_ewa_requests(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["EWA_Request_ID"]),
            text(row["Opportunity_ID"]),
            text(row["Opportunity_Role_ID"]),
            text(row["Employee_ID"]),
            text(row["RequestType"]),
            text(row["EWAStatus"]),
            as_float(row["RequestedFTE"]),
            text(row["ProposedStartDate"]),
            text(row["ProposedEndDate"]),
            as_bool(row["ApprovalRequired"]),
            text(row["BookingOwner"]),
            text(row["BlockingReason"]),
            text(row["NextAction"]),
            text(row["LastUpdated"]),
            text(row["Notes"]),
            as_float(row["AvailableFTEAtStart"]),
            as_float(row["FTEGap"]),
            as_bool(row["CanSplitRole"]),
            text(row["EarliestFullAvailabilityDate"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "EwaRequest"
          ("id", "opportunityId", "opportunityRoleId", "personId", "requestType",
           "ewaStatus", "requestedFte", "proposedStartDate", "proposedEndDate",
           "approvalRequired", "bookingOwner", "blockingReason", "nextAction",
           "lastUpdated", "notes", "availableFteAtStart", "fteGap", "canSplitRole",
           "earliestFullAvailabilityDate")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def insert_scenario_targets(
    connection: sqlite3.Connection, rows: list[dict[str, Any]]
) -> int:
    payload = [
        (
            text(row["Scenario_ID"]),
            text(row["ScenarioName"]),
            text(row["TargetDate"]),
            as_float(row["TargetBenchRate"]),
            as_int(row["TargetBenchHeadcount"]),
            text(row["Focus"]),
            text(row["SuccessMeasure"]),
        )
        for row in rows
    ]
    connection.executemany(
        """
        INSERT INTO "ScenarioTarget"
          ("id", "scenarioName", "targetDate", "targetBenchRate", "targetBenchHeadcount",
           "focus", "successMeasure")
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def main() -> None:
    args = parse_args()
    excel_path = args.excel.resolve()
    db_path = args.db.resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheets = {sheet_name: read_sheet(workbook, sheet_name) for sheet_name in workbook.sheetnames}
    ensure_required_sheets(sheets)

    connection = create_database(db_path, replace=args.replace)
    extend_schema(connection)
    finalize_schema(connection)

    counts: dict[str, int] = {}
    try:
        with connection:
            batch_id = insert_import_batch(
                connection,
                workbook_name=excel_path.name,
                version=workbook_version(sheets),
            )
            counts["ImportBatch"] = 1
            counts["RawSheetRow"] = load_raw_sheet_rows(connection, batch_id, sheets)

            people_rows = [row["values"] for row in sheets["People"]["rows"]]
            skill_rows = [row["values"] for row in sheets["Skills"]["rows"]]
            skill_catalog_rows = [row["values"] for row in sheets["Skill Catalog"]["rows"]]
            profile_rows = [row["values"] for row in sheets["Profiles"]["rows"]]
            allocation_rows = [row["values"] for row in sheets["Allocations"]["rows"]]
            bench_rows = [row["values"] for row in sheets["Bench"]["rows"]]
            partial_rows = [row["values"] for row in sheets["Partial Capacity"]["rows"]]
            availability_rows = [row["values"] for row in sheets["Availability Calendar"]["rows"]]
            bench_movement_rows = [row["values"] for row in sheets["Bench Movement"]["rows"]]
            history_rows = [row["values"] for row in sheets["Project History"]["rows"]]
            opportunity_rows = [row["values"] for row in sheets["Opportunities"]["rows"]]
            opportunity_role_rows = [row["values"] for row in sheets["Opportunity Roles"]["rows"]]
            overlay_rows = [row["values"] for row in sheets["Opportunity Overlays"]["rows"]]
            ewa_rows = [row["values"] for row in sheets["EWA Requests"]["rows"]]
            scenario_rows = [row["values"] for row in sheets["Scenario Targets"]["rows"]]

            counts["Person"] = insert_people(connection, people_rows)
            counts["PersonAvailabilitySnapshot"] = insert_person_snapshots(connection, people_rows)
            counts["Profile"] = insert_profiles(connection, profile_rows)
            counts["SkillCatalog"] = insert_skill_catalog(connection, skill_catalog_rows)
            counts["PersonSkillEvidence"] = insert_person_skills(connection, skill_rows)
            counts["CurrentAllocation"] = insert_current_allocations(connection, allocation_rows)
            counts["SupplyRecord"] = insert_supply_records(connection, bench_rows)
            counts["PartialCapacityView"] = insert_partial_capacity_view(connection, partial_rows)
            counts["AvailabilityWeek"] = insert_availability_weeks(connection, availability_rows)
            counts["BenchMovementWeek"] = insert_bench_movement(connection, bench_movement_rows)
            counts["ProjectHistory"] = insert_project_history(connection, history_rows)
            counts["Opportunity"] = insert_opportunities(connection, opportunity_rows)

            role_count, requirement_count = insert_opportunity_roles(
                connection, opportunity_role_rows
            )
            counts["OpportunityRole"] = role_count
            counts["OpportunityRoleSkillRequirement"] = requirement_count
            counts["OpportunityCandidateOverlay"] = insert_overlays(connection, overlay_rows)
            counts["EwaRequest"] = insert_ewa_requests(connection, ewa_rows)
            counts["ScenarioTarget"] = insert_scenario_targets(connection, scenario_rows)
    finally:
        connection.close()

    print(f"Created SQLite database: {db_path}")
    print(f"Imported workbook: {excel_path.name}")
    for table_name in sorted(counts):
        print(f"{table_name}: {counts[table_name]}")


if __name__ == "__main__":
    main()
